"""
المسارات الخاصة بالموزعين
"""
from flask import Blueprint, render_template, jsonify, request, session, redirect, url_for, flash
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timedelta
from models import SupportTicket, db, Reseller, User, ActivationCode, Device, DeviceActivationCode
import uuid
import re
from flask import send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from audit_helper import log_reseller_action, log_user_action

reseller_bp = Blueprint('reseller', __name__)

# ============================================================================
# 🟢 دوال مساعدة لحساب الإحصائيات
# ============================================================================

def get_dashboard_stats(reseller_id):
    """
    حساب إحصائيات لوحة التحكم للموزع
    """
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = today_start - timedelta(days=1)
    
    # 🔴 إجمالي الاشتراكات (التفعيلات)
    total_activations = ActivationCode.query.filter_by(
        reseller_id=reseller_id
    ).count()
    
    # 🔴 الاشتراكات اليوم
    today_activations = ActivationCode.query.filter(
        ActivationCode.reseller_id == reseller_id,
        ActivationCode.created_at >= today_start
    ).count()
    
    # 🔴 الاشتراكات أمس
    yesterday_activations = ActivationCode.query.filter(
        ActivationCode.reseller_id == reseller_id,
        ActivationCode.created_at >= yesterday_start,
        ActivationCode.created_at < today_start
    ).count()
    
    # حساب نسبة التغير اليومي
    daily_change_percent = 0
    if yesterday_activations > 0:
        daily_change_percent = ((today_activations - yesterday_activations) / yesterday_activations) * 100
    elif today_activations > 0:
        daily_change_percent = 100  # زيادة من 0 إلى رقم موجب
    
    # 🔴 الاشتراكات المنتهية الصلاحية
    expired_subscriptions = ActivationCode.query.filter(
        ActivationCode.reseller_id == reseller_id,
        ActivationCode.expiration_date < now
    ).count()
    
    # 🔴 الاشتراكات النشطة (المفعلة وغير منتهية)
    active_subscriptions = ActivationCode.query.filter(
        ActivationCode.reseller_id == reseller_id,
        ActivationCode.activated_at != None,
        ActivationCode.expiration_date >= now
    ).count()
    
    # حساب نسبة التغير للاشتراكات النشطة
    # (مقارنة مع أسبوع ماضي)
    week_ago = now - timedelta(days=7)
    active_week_ago = ActivationCode.query.filter(
        ActivationCode.reseller_id == reseller_id,
        ActivationCode.activated_at != None,
        ActivationCode.created_at <= week_ago
    ).count()
    
    active_change_percent = 0
    if active_week_ago > 0:
        active_change_percent = ((active_subscriptions - active_week_ago) / active_week_ago) * 100
    elif active_subscriptions > 0:
        active_change_percent = 100
    
    return {
        'total_activations': total_activations,
        'today_activations': today_activations,
        'daily_change_percent': round(daily_change_percent, 1),
        'active_subscriptions': active_subscriptions,
        'active_change_percent': round(active_change_percent, 1),
        'expired_subscriptions': expired_subscriptions
    }

@reseller_bp.route('/dashboard')
def dashboard():
    """صفحة لوحة تحكم الموزع"""
    if 'reseller_id' not in session:
        return redirect(url_for('reseller.login'))
    
    # الحصول على بيانات الموزع من قاعدة البيانات
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        session.clear()
        return redirect(url_for('reseller.login'))
    
    # حساب الإحصائيات
    stats = get_dashboard_stats(reseller.id)
    
    return render_template('reseller/dashboard.html', reseller=reseller, stats=stats)



@reseller_bp.route('/activatecode')
def activate_code():
    """صفحة لوحة تحكم الموزع"""
    if 'reseller_id' not in session:
        return redirect(url_for('reseller.login'))
    
    # الحصول على بيانات الموزع من قاعدة البيانات
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        session.clear()
        return redirect(url_for('reseller.login'))
    
    return render_template('reseller/activate-code.html', reseller=reseller)



@reseller_bp.route('/login', methods=['GET', 'POST'])
def login():
    """صفحة تسجيل دخول الموزع"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        # التحقق من البيانات
        if not email or not password:
            flash('Please enter email and password', 'error')
            return render_template('reseller/login.html')
        
        # البحث عن الموزع
        reseller = Reseller.query.filter_by(email=email).first()
        
        if reseller and check_password_hash(reseller.password_hash, password):
            if not reseller.is_active:
                flash('Your account is inactive. Contact support.', 'error')
                # تسجيل محاولة دخول فاشلة
                log_reseller_action(
                    reseller_id=reseller.id,
                    action='login',
                    description=f'Failed login attempt - account is inactive'
                )
                return render_template('reseller/login.html')
            
            # تعيين الجلسة
            session['reseller_id'] = reseller.id
            session['reseller_name'] = reseller.name
            session['reseller_email'] = reseller.email
            
            # تسجيل عملية الدخول الناجحة
            log_reseller_action(
                reseller_id=reseller.id,
                action='login',
                description=f'Reseller {reseller.name} logged in successfully'
            )
            
            # التحقق: هل لدى الموزع PIN؟
            if not reseller.pin_hash:
                # لا يوجد PIN، توجيه لصفحة تعيين PIN
                flash('Please set your PIN code', 'info')
                return redirect(url_for('reseller.setup_pin'))
            
            flash(f'Welcome back, {reseller.name}!', 'success')
            return redirect(url_for('reseller.dashboard'))
        else:
            flash('Invalid email or password', 'error')
            return render_template('reseller/login.html')
    
    return render_template('reseller/login.html')

@reseller_bp.route('/logout', methods=['POST'])
def logout():
    """تسجيل خروج الموزع"""
    reseller_id = session.get('reseller_id')
    if reseller_id:
        log_reseller_action(
            reseller_id=reseller_id,
            action='logout',
            description=f'Reseller logged out'
        )
    session.clear()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('reseller.login'))

@reseller_bp.route('/setup-pin', methods=['GET', 'POST'])
def setup_pin():
    """صفحة تعيين PIN للموزع"""
    if 'reseller_id' not in session:
        return redirect(url_for('reseller.login'))
    
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        session.clear()
        return redirect(url_for('reseller.login'))
    
    # إذا كان لديه PIN بالفعل، توجيه للـ dashboard
    if reseller.pin_hash:
        return redirect(url_for('reseller.dashboard'))
    
    if request.method == 'POST':
        pin = request.form.get('pin', '').strip()
        pin_confirm = request.form.get('pin_confirm', '').strip()
        
        # التحقق من صحة البيانات
        if not pin or not pin_confirm:
            flash('Please enter PIN code', 'error')
            return render_template('reseller/setup_pin.html', reseller=reseller)
        
        # التحقق: يجب أن يكون 4 أرقام
        if not pin.isdigit() or len(pin) != 4:
            flash('PIN must be exactly 4 digits', 'error')
            return render_template('reseller/setup_pin.html', reseller=reseller)
        
        # التحقق: تطابق PIN والتأكيد
        if pin != pin_confirm:
            flash('PIN codes do not match', 'error')
            return render_template('reseller/setup_pin.html', reseller=reseller)
        
        try:
            # حفظ PIN المشفر
            reseller.pin_hash = generate_password_hash(pin)
            db.session.commit()
            
            flash('PIN set successfully! You can now access your dashboard.', 'success')
            return redirect(url_for('reseller.dashboard'))
        except Exception as e:
            db.session.rollback()
            flash('Error setting PIN. Please try again.', 'error')
            return render_template('reseller/setup_pin.html', reseller=reseller)
    
    return render_template('reseller/setup_pin.html', reseller=reseller)

@reseller_bp.route('/api/my-codes', methods=['GET'])
def get_activation_codes():
    """الحصول على أكواد التفعيل الخاصة بالموزع"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    
    # الحصول على جميع الأكواز الخاصة بهذا الموزع
    activation_codes = ActivationCode.query.filter_by(
        reseller_id=reseller_id
    ).order_by(ActivationCode.created_at.desc()).all()
    
    codes_data = []
    for code in activation_codes:
        # الحصول على بيانات المستخدم إن وجدت
        user = User.query.get(code.assigned_user_id) if code.assigned_user_id else None
        
        # حساب الحالة
        now = datetime.utcnow()
        if code.expiration_date and code.expiration_date < now:
            status = "Expired"
        elif code.activated_at:
            status = "Active"
        else:
            status = "Not Activated"
        
        codes_data.append({
            'id': code.id,
            'code': code.code,
            'username': user.username if user else 'N/A',
            'duration_months': code.duration_months,
            'max_devices': code.max_devices,
            'is_lifetime': code.is_lifetime,
            'status': status,
            'activated_at': code.activated_at.isoformat() if code.activated_at else None,
            'expiration_date': code.expiration_date.isoformat() if code.expiration_date else None,
            'created_at': code.created_at.isoformat() if code.created_at else None,
        })
    
    return jsonify({
        'success': True,
        'data': codes_data,
        'count': len(codes_data)
    }), 200

@reseller_bp.route('/api/my-users', methods=['GET'])
def get_users():
    """الحصول على المستخدمين الخاصين بالموزع"""
    return jsonify([])


@reseller_bp.route('/api/profile', methods=['GET'])
def get_profile():
    """الحصول على بيانات الموزع الحالي"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        return jsonify({'success': False, 'message': 'Reseller not found'}), 404
    
    return jsonify({
        'success': True,
        'data': {
            'id': reseller.id,
            'name': reseller.name,
            'email': reseller.email,
            'country': reseller.country,
            'points_balance': reseller.points_balance,
            'is_active': reseller.is_active,
            'total_amount_charged': reseller.total_amount_charged,
            'total_points_charged': reseller.total_points_charged,
            'created_at': reseller.created_at.isoformat() if reseller.created_at else None
        }
    }), 200


@reseller_bp.route('/api/verify-pin', methods=['POST'])
def verify_pin():
    """التحقق من PIN الموزع"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        return jsonify({'success': False, 'message': 'Reseller not found'}), 404
    
    data = request.get_json()
    pin = data.get('pin', '').strip()
    
    # التحقق من صحة البيانات
    if not pin or len(pin) != 4 or not pin.isdigit():
        return jsonify({'success': False, 'message': 'Invalid PIN format'}), 400
    
    # التحقق من PIN
    if reseller.pin_hash and check_password_hash(reseller.pin_hash, pin):
        return jsonify({'success': True, 'message': 'PIN verified successfully'}), 200
    else:
        return jsonify({'success': False, 'message': 'Invalid PIN'}), 401


@reseller_bp.route('/api/export-codes', methods=['GET'])
def export_codes():
    """تصدير أكواد التفعيل إلى Excel بصيغة احترافية"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    reseller = Reseller.query.get(reseller_id)
    
    # الحصول على جميع الأكواز الخاصة بهذا الموزع
    activation_codes = ActivationCode.query.filter_by(
        reseller_id=reseller_id
    ).order_by(ActivationCode.created_at.desc()).all()
    
    # إنشاء Workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "Activation Codes"
    
    # تعريف الأنماط
    # Header Style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border Style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Data Style
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_font = Font(size=10, color="000000")
    
    # Row alternating colors
    light_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status colors
    active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    active_font = Font(color="006100", bold=True)
    
    expired_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    expired_font = Font(color="9C0006", bold=True)
    
    not_activated_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    not_activated_font = Font(color="9C6500", bold=True)
    
    # Headers
    headers = [
        'Activation Code',
        'Username',
        'Plan Type',
        'Max Devices',
        'Status',
        'Created Date',
        'Expiration Date',
        'Created At (Full)'
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set header row height
    ws.row_dimensions[1].height = 25
    
    # Write data
    now = datetime.utcnow()
    for row_idx, code in enumerate(activation_codes, 2):
        user = User.query.get(code.assigned_user_id) if code.assigned_user_id else None
        
        # حساب الحالة
        if code.expiration_date and code.expiration_date < now:
            status = "Expired"
        elif code.activated_at:
            status = "Active"
        else:
            status = "Not Activated"
        
        # تحديد نوع الخطة
        plan_type = "Lifetime" if code.is_lifetime else "1 Year"
        
        # تنسيق التواريخ
        created_date = code.created_at.strftime('%Y-%m-%d') if code.created_at else 'N/A'
        expiration_date = code.expiration_date.strftime('%Y-%m-%d') if code.expiration_date else 'N/A'
        created_at_full = code.created_at.strftime('%Y-%m-%d %H:%M:%S') if code.created_at else 'N/A'
        
        row_data = [
            code.code,
            user.username if user else 'N/A',
            plan_type,
            code.max_devices,
            status,
            created_date,
            expiration_date,
            created_at_full
        ]
        
        # اختيار لون الصف بناءً على الحالة
        row_fill = light_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = data_alignment
            cell.font = data_font
            
            # تطبيق ألوان الحالة على عمود Status فقط
            if col_idx == 5:  # Status column
                if status == "Active":
                    cell.fill = active_fill
                    cell.font = active_font
                elif status == "Expired":
                    cell.fill = expired_fill
                    cell.font = expired_font
                elif status == "Not Activated":
                    cell.fill = not_activated_fill
                    cell.font = not_activated_font
            else:
                cell.fill = row_fill
        
        # Set row height
        ws.row_dimensions[row_idx].height = 20
    
    # ضبط عرض الأعمدة
    column_widths = {
        'A': 20,  # Activation Code
        'B': 18,  # Username
        'C': 15,  # Plan Type
        'D': 12,  # Max Devices
        'E': 16,  # Status
        'F': 15,  # Created Date
        'G': 16,  # Expiration Date
        'H': 22   # Created At (Full)
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # إضافة صف معلومات في الأعلى
    ws.insert_rows(1)
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Activation Codes Report - {reseller.name}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    
    # إضافة معلومات التصدير
    ws.insert_rows(2)
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f"Exported on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(activation_codes)}"
    info_cell.font = Font(italic=True, size=9, color="666666")
    info_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    info_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    
    # إرجاع الملف
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)
    
    filename = f"activation_codes_{reseller.name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        mem,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================================
# 🔴 المرحلة 3-8: نظام تفعيل الأكواز
# ============================================================================

def generate_username(base_name=None):
    """توليد اسم مستخدم فريد"""
    if base_name:
        base_name = re.sub(r'[^a-zA-Z0-9_]', '', base_name)
        if base_name:
            # تحقق من أن الاسم غير موجود
            existing = User.query.filter_by(username=base_name).first()
            if not existing:
                return base_name
    
    # إنشاء اسم عشوائي
    while True:
        random_user = f"SERVO-{uuid.uuid4().hex[:8]}"
        if not User.query.filter_by(username=random_user).first():
            return random_user


@reseller_bp.route('/api/activate-code', methods=['POST'])
def activate_code_api():
    """
    🔴 المراحل 3-8: نظام تفعيل الأكواز الكامل
    
    متطلبات:
    - activation_code (مطلوب)
    - subscription_duration (مطلوب): '1year' أو 'lifetime'
    - username (اختياري): اسم المستخدم
    - media_link (اختياري): رابط الوسائط
    
    ملاحظة: 
    - عدد الأجهزة ثابت = 1
    - خصم النقاط: 1 نقطة للسنة الواحدة، نقطتان لمدى الحياة
    """
    
    # التحقق من تسجيل دخول الموزع
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    reseller = Reseller.query.get(reseller_id)
    if not reseller:
        return jsonify({'success': False, 'message': 'Reseller not found'}), 404
    
    # الحصول على البيانات من الطلب
    data = request.get_json()
    
    activation_code = data.get('activationCode', '').strip()
    subscription_duration = data.get('subscriptionDuration', '').strip()  # '1year' or 'lifetime'
    username = data.get('username', '').strip() if data.get('username') else None
    media_link = data.get('mediaLink', '').strip() if data.get('mediaLink') else None
    
    # ============================================================================
    # 🟢 التحقق من صحة البيانات
    # ============================================================================
    if not activation_code:
        return jsonify({'success': False, 'message': 'Activation code is required'}), 400
    
    if subscription_duration not in ['1year', 'lifetime']:
        return jsonify({'success': False, 'message': 'Invalid subscription duration. Must be "1year" or "lifetime"'}), 400
    
    # ============================================================================
    # 🟢 المرحلة 4: التحقق من كود التفعيل
    # ============================================================================
    
    # البحث عن الكود في device_activation_codes
    device_activation_code = DeviceActivationCode.query.filter_by(
        activation_code=activation_code
    ).first()
    
    # التحقق: غير موجود
    if not device_activation_code:
        return jsonify({'success': False, 'message': '❌ Activation code not found'}), 404
    
    # التحقق: مستخدم من قبل
    if device_activation_code.is_used:
        return jsonify({'success': False, 'message': '❌ Activation code has already been used'}), 400
    
    # التحقق: منتهي الصلاحية
    if device_activation_code.expires_at and datetime.utcnow() > device_activation_code.expires_at:
        return jsonify({'success': False, 'message': '❌ Activation code has expired'}), 400
    
    # ============================================================================
    # 🟢 حساب مدة الاشتراك وخصم النقاط
    # ============================================================================
    
    is_lifetime = subscription_duration == 'lifetime'
    duration_months = 12 if subscription_duration == '1year' else 120  # 120 شهر لمدى الحياة (10 سنوات)
    
    # حساب خصم النقاط
    points_to_deduct = 2 if is_lifetime else 1
    
    # التحقق من أن الموزع لديه نقاط كافية
    if reseller.points_balance < points_to_deduct:
        return jsonify({
            'success': False, 
            'message': f'❌ Insufficient points. Required: {points_to_deduct}, Available: {reseller.points_balance}'
        }), 400
    
    # ============================================================================
    # 🟢 المرحلة 5: إنشاء المستخدم
    # ============================================================================
    
    # توليد أو استخدام اسم المستخدم المعطى
    final_username = generate_username(username)
    
    try:
        # إنشاء سجل جديد في جدول users
        new_user = User(
            username=final_username,
            reseller_id=reseller_id
        )
        db.session.add(new_user)
        db.session.flush()  # للحصول على user_id قبل الـ commit
        user_id = new_user.id
        
        # ============================================================================
        # 🟢 المرحلة 6: تفعيل الاشتراك
        # ============================================================================
        
        # حساب تاريخ انتهاء الاشتراك
        if is_lifetime:
            # لمدى الحياة: لا يوجد تاريخ انتهاء (أو سنة 2099)
            expiration_date = datetime.utcnow() + timedelta(days=365*100)  # 100 سنة
        else:
            # سنة واحدة: 12 شهر
            expiration_date = datetime.utcnow() + timedelta(days=365)
        
        new_activation_code = ActivationCode(
            code=activation_code,
            reseller_id=reseller_id,
            assigned_user_id=user_id,
            duration_months=duration_months,
            max_devices=1,  # ثابت دائماً = 1
            is_lifetime=is_lifetime,
            activated_at=datetime.utcnow(),
            expiration_date=expiration_date
        )
        db.session.add(new_activation_code)
        
        # ============================================================================
        # 🟢 المرحلة 7: ربط الجهاز بالمستخدم
        # ============================================================================
        
        # استخدام device_id الموجود من التسجيل الأولي للجهاز
        device_uid = device_activation_code.device_id
        
        new_device = Device(
            user_id=user_id,
            device_uid=device_uid,
            device_type=device_activation_code.device_type or 'unknown',
            is_active=True,
            first_login_at=datetime.utcnow(),
            media_link=media_link
        )
        db.session.add(new_device)
        
        # ============================================================================
        # 🟢 المرحلة 8: خصم النقاط والإغلاق
        # ============================================================================
        
        # خصم النقاط من رصيد الموزع
        reseller.points_balance -= points_to_deduct
        reseller.total_points_charged += points_to_deduct
        
        # إغلاق كود التفعيل
        device_activation_code.is_used = True
        device_activation_code.used_at = datetime.utcnow()
        device_activation_code.activated_by_reseller_id = reseller_id
        device_activation_code.user_id = user_id
        device_activation_code.username = final_username
        
        # حفظ جميع التغييرات
        db.session.commit()
        
        # تسجيل العملية
        log_reseller_action(
            reseller_id=reseller_id,
            action='activate',
            description=f'Activation code {activation_code} was activated for user {final_username}. Points deducted: {points_to_deduct}',
            resource_type='activation_code',
            resource_id=new_activation_code.id
        )
        
        return jsonify({
            'success': True,
            'message': '✅ Activation code processed successfully',
            'data': {
                'user_id': user_id,
                'username': final_username,
                'device_uid': device_uid,
                'is_lifetime': is_lifetime,
                'subscription_duration': subscription_duration,
                'expiration_date': expiration_date.isoformat(),
                'max_devices': 1,
                'points_deducted': points_to_deduct,
                'remaining_points': reseller.points_balance
            }
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error processing activation code: {str(e)}'
        }), 500


# ============================================================================
# 📊 API لجلب بيانات التحليلات
# ============================================================================

@reseller_bp.route('/api/analytics', methods=['GET'])
def get_analytics():
    """جلب بيانات التحليلات حسب الفترة الزمنية"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    period = request.args.get('period', 'daily')  # daily, weekly, monthly, yearly
    
    now = datetime.utcnow()
    
    try:
        if period == 'daily':
            # آخر 7 أيام
            data = get_daily_analytics(reseller_id, now)
        elif period == 'weekly':
            # آخر 4 أسابيع
            data = get_weekly_analytics(reseller_id, now)
        elif period == 'monthly':
            # آخر 6 أشهر
            data = get_monthly_analytics(reseller_id, now)
        elif period == 'yearly':
            # آخر 5 سنوات
            data = get_yearly_analytics(reseller_id, now)
        else:
            return jsonify({'success': False, 'message': 'Invalid period'}), 400
        
        return jsonify({'success': True, 'data': data}), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def get_daily_analytics(reseller_id, now):
    """حساب بيانات آخر 7 أيام"""
    labels = []
    activations_data = []
    
    days_names = ['الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']
    
    for i in range(6, -1, -1):  # من 6 أيام ماضية إلى اليوم
        date = now - timedelta(days=i)
        day_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        # عدد التفعيلات في هذا اليوم
        activations = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.created_at >= day_start,
            ActivationCode.created_at < day_end
        ).count()
        
        # اسم اليوم بالعربية
        day_name = days_names[date.weekday()]
        
        labels.append(day_name)
        activations_data.append(activations)
    
    # حساب النسبة المئوية للنمو
    if activations_data[-2] > 0:
        trend = ((activations_data[-1] - activations_data[-2]) / activations_data[-2]) * 100
    elif activations_data[-1] > 0:
        trend = 100
    else:
        trend = 0
    
    return {
        'labels': labels,
        'activations': activations_data,
        'trend': round(trend, 1)
    }


def get_weekly_analytics(reseller_id, now):
    """حساب بيانات آخر 4 أسابيع"""
    labels = []
    activations_data = []
    
    for i in range(3, -1, -1):  # من 3 أسابيع ماضية إلى الأسبوع الحالي
        week_start = (now - timedelta(weeks=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        # ابدأ من بداية الأسبوع (الأحد)
        week_start = week_start - timedelta(days=week_start.weekday() + 1)
        week_end = week_start + timedelta(days=7)
        
        # عدد التفعيلات في هذا الأسبوع
        activations = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.created_at >= week_start,
            ActivationCode.created_at < week_end
        ).count()
        
        # اسم الأسبوع
        label = f'الأسبوع {i + 1}'
        
        labels.append(label)
        activations_data.append(activations)
    
    # حساب النسبة المئوية للنمو
    if activations_data[-2] > 0:
        trend = ((activations_data[-1] - activations_data[-2]) / activations_data[-2]) * 100
    elif activations_data[-1] > 0:
        trend = 100
    else:
        trend = 0
    
    return {
        'labels': labels,
        'activations': activations_data,
        'trend': round(trend, 1)
    }


def get_monthly_analytics(reseller_id, now):
    """حساب بيانات آخر 6 أشهر"""
    labels = []
    activations_data = []
    
    months_names = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 
                    'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    
    for i in range(5, -1, -1):  # من 5 أشهر ماضية إلى الشهر الحالي
        if i == 0:
            # الشهر الحالي
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            # الأشهر الماضية
            month_date = now - timedelta(days=30*i)
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # حساب بداية الشهر التالي
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        
        # عدد التفعيلات في هذا الشهر
        activations = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.created_at >= month_start,
            ActivationCode.created_at < month_end
        ).count()
        
        # اسم الشهر
        month_name = months_names[month_start.month - 1]
        
        labels.append(month_name)
        activations_data.append(activations)
    
    # حساب النسبة المئوية للنمو
    if activations_data[-2] > 0:
        trend = ((activations_data[-1] - activations_data[-2]) / activations_data[-2]) * 100
    elif activations_data[-1] > 0:
        trend = 100
    else:
        trend = 0
    
    return {
        'labels': labels,
        'activations': activations_data,
        'trend': round(trend, 1)
    }


def get_yearly_analytics(reseller_id, now):
    """حساب بيانات آخر 5 سنوات"""
    labels = []
    activations_data = []
    
    for i in range(4, -1, -1):  # من 4 سنوات ماضية إلى السنة الحالية
        year = now.year - i
        year_start = now.replace(year=year, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        year_end = year_start.replace(year=year + 1)
        
        # عدد التفعيلات في هذه السنة
        activations = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.created_at >= year_start,
            ActivationCode.created_at < year_end
        ).count()
        
        labels.append(str(year))
        activations_data.append(activations)
    
    # حساب النسبة المئوية للنمو
    if activations_data[-2] > 0:
        trend = ((activations_data[-1] - activations_data[-2]) / activations_data[-2]) * 100
    elif activations_data[-1] > 0:
        trend = 100
    else:
        trend = 0
    
    return {
        'labels': labels,
        'activations': activations_data,
        'trend': round(trend, 1)
    }


# ============================================================================
# 📊 API لجلب نسبة الاشتراكات السنوية والمدى الحياة
# ============================================================================

@reseller_bp.route('/api/subscription-types', methods=['GET'])
def get_subscription_types():
    """جلب نسبة الاشتراكات السنوية والمدى الحياة"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    
    try:
        # عدد الاشتراكات السنوية
        yearly_subs = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.is_lifetime == False
        ).count()
        
        # عدد الاشتراكات المدى الحياة
        lifetime_subs = ActivationCode.query.filter(
            ActivationCode.reseller_id == reseller_id,
            ActivationCode.is_lifetime == True
        ).count()
        
        total = yearly_subs + lifetime_subs
        
        if total == 0:
            yearly_percent = 0
            lifetime_percent = 0
        else:
            yearly_percent = (yearly_subs / total) * 100
            lifetime_percent = (lifetime_subs / total) * 100
        
        return jsonify({
            'success': True,
            'data': {
                'yearly': yearly_subs,
                'lifetime': lifetime_subs,
                'total': total,
                'yearly_percent': round(yearly_percent, 1),
                'lifetime_percent': round(lifetime_percent, 1)
            }
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500



# ============================================================================
# ============================================================================

@reseller_bp.route('/support', methods=['GET'])
def support():
    """صفحة الدعم الفني للموزعين"""
    if 'reseller_id' not in session:
        return redirect(url_for('reseller.login'))
    
    reseller = Reseller.query.get(session['reseller_id'])
    if not reseller:
        session.clear()
        return redirect(url_for('reseller.login'))
    
    # جلب جميع تذاكر الموزع
    tickets = SupportTicket.query.filter_by(reseller_id=reseller.id).order_by(SupportTicket.created_at.desc()).all()
    
    # تحضير بيانات التذاكر
    tickets_data = []
    for ticket in tickets:
        messages = []
        if ticket.messages:
            for msg in ticket.messages:
                messages.append({
                    'id': msg.id,
                    'sender_type': msg.sender_type,
                    'sender_id': msg.sender_id,
                    'message': msg.message,
                    'is_internal': msg.is_internal,
                    'created_at': msg.created_at.isoformat() if msg.created_at else None
                })
        
        tickets_data.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'subject': ticket.subject,
            'description': ticket.description,
            'priority': ticket.priority,
            'status': ticket.status,
            'reseller_id': ticket.reseller_id,
            'created_at': ticket.created_at.isoformat() if ticket.created_at else None,
            'updated_at': ticket.updated_at.isoformat() if ticket.updated_at else None,
            'resolved_at': ticket.resolved_at.isoformat() if ticket.resolved_at else None,
            'messages': messages,
            'message_count': len(messages)
        })
    
    return render_template('reseller/support.html', reseller=reseller, tickets=tickets_data)


# ============================================================================
# 🎫 Support Tickets API
# ============================================================================

def generate_ticket_number():
    """توليد رقم تذكرة فريد"""
    from datetime import datetime
    timestamp = int(datetime.utcnow().timestamp())
    last_ticket = db.session.query(db.func.max(db.cast(
        db.func.substr(SupportTicket.ticket_number, 3), db.Integer
    ))).scalar()
    
    next_number = (last_ticket or 0) + 1
    return f"T-{next_number:06d}"


@reseller_bp.route('/api/tickets/create', methods=['POST'])
def create_ticket():
    """إنشاء تذكرة دعم جديدة"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    reseller = Reseller.query.get(reseller_id)
    
    if not reseller:
        return jsonify({'success': False, 'message': 'Reseller not found'}), 404
    
    # الحصول على البيانات
    data = request.get_json()
    subject = data.get('subject', '').strip()
    description = data.get('description', '').strip()
    priority = data.get('priority', 'normal').strip()
    
    # التحقق من صحة البيانات
    if not subject or len(subject) < 5:
        return jsonify({'success': False, 'message': 'Subject must be at least 5 characters'}), 400
    
    if not description or len(description) < 10:
        return jsonify({'success': False, 'message': 'Description must be at least 10 characters'}), 400
    
    if priority not in ['low', 'normal', 'high', 'urgent']:
        priority = 'normal'
    
    try:
        from models import SupportTicket
        
        # إنشاء التذكرة
        ticket_number = generate_ticket_number()
        
        ticket = SupportTicket(
            ticket_number=ticket_number,
            reseller_id=reseller_id,
            subject=subject,
            description=description,
            priority=priority,
            status='open'
        )
        
        db.session.add(ticket)
        db.session.commit()
        
        # تسجيل العملية
        log_reseller_action(
            reseller_id=reseller_id,
            action='create_ticket',
            description=f'Created support ticket {ticket_number}: {subject}',
            resource_type='support_ticket',
            resource_id=ticket.id
        )
        
        return jsonify({
            'success': True,
            'message': 'Ticket created successfully',
            'ticket_number': ticket_number,
            'ticket_id': ticket.id
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@reseller_bp.route('/api/tickets', methods=['GET'])
def get_tickets():
    """الحصول على جميع تذاكر الدعم الخاصة بالموزع"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    
    try:
        from models import SupportTicket
        
        tickets = SupportTicket.query.filter_by(reseller_id=reseller_id)\
            .order_by(SupportTicket.created_at.desc()).all()
        
        tickets_data = []
        for ticket in tickets:
            tickets_data.append({
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'subject': ticket.subject,
                'description': ticket.description,
                'priority': ticket.priority,
                'status': ticket.status,
                'created_at': ticket.created_at.isoformat() if ticket.created_at else None,
                'updated_at': ticket.updated_at.isoformat() if ticket.updated_at else None,
                'resolved_at': ticket.resolved_at.isoformat() if ticket.resolved_at else None,
                'message_count': len(ticket.messages) if ticket.messages else 0
            })
        
        return jsonify({
            'success': True,
            'data': tickets_data,
            'count': len(tickets_data)
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@reseller_bp.route('/api/tickets/<int:ticket_id>', methods=['GET'])
def get_ticket(ticket_id):
    """الحصول على تفاصيل تذكرة دعم معينة"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    
    try:
        from models import SupportTicket
        
        ticket = SupportTicket.query.filter_by(
            id=ticket_id,
            reseller_id=reseller_id
        ).first()
        
        if not ticket:
            return jsonify({'success': False, 'message': 'Ticket not found'}), 404
        
        messages = []
        for msg in ticket.messages:
            messages.append({
                'id': msg.id,
                'sender_type': msg.sender_type,
                'sender_id': msg.sender_id,
                'message': msg.message,
                'is_internal': msg.is_internal,
                'created_at': msg.created_at.isoformat() if msg.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': {
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'subject': ticket.subject,
                'description': ticket.description,
                'priority': ticket.priority,
                'status': ticket.status,
                'created_at': ticket.created_at.isoformat() if ticket.created_at else None,
                'updated_at': ticket.updated_at.isoformat() if ticket.updated_at else None,
                'resolved_at': ticket.resolved_at.isoformat() if ticket.resolved_at else None,
                'messages': messages
            }
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@reseller_bp.route('/api/tickets/<int:ticket_id>/message', methods=['POST'])
def add_ticket_message(ticket_id):
    """إضافة رسالة إلى تذكرة دعم"""
    if 'reseller_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    reseller_id = session['reseller_id']
    reseller = Reseller.query.get(reseller_id)
    
    if not reseller:
        return jsonify({'success': False, 'message': 'Reseller not found'}), 404
    
    try:
        from models import SupportTicket, TicketMessage
        
        ticket = SupportTicket.query.filter_by(
            id=ticket_id,
            reseller_id=reseller_id
        ).first()
        
        if not ticket:
            return jsonify({'success': False, 'message': 'Ticket not found'}), 404
        
        data = request.get_json()
        message_text = data.get('message', '').strip()
        
        if not message_text:
            return jsonify({'success': False, 'message': 'Message cannot be empty'}), 400
        
        # إنشاء الرسالة
        message = TicketMessage(
            ticket_id=ticket_id,
            sender_type='reseller',
            sender_id=reseller_id,
            message=message_text,
            is_internal=False
        )
        
        db.session.add(message)
        db.session.commit()
        
        # تحديث تاريخ آخر تحديث للتذكرة
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Message added successfully',
            'data': {
                'id': message.id,
                'sender_type': message.sender_type,
                'sender_id': message.sender_id,
                'message': message.message,
                'created_at': message.created_at.isoformat() if message.created_at else None
            }
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500